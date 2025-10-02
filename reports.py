# reports.py
import os
import sqlite3
from datetime import datetime, timedelta

from PySide6.QtCore import QDate
from PySide6.QtWidgets import (
    QDateEdit,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from db import connect as _connect


# choose invoice date column that exists in your DB
def _invoice_date_expr(cur) -> str:
    cur.execute("PRAGMA table_info(invoices)")
    cols = [row[1] for row in cur.fetchall()]
    return "DATE(i.invoice_date)" if "invoice_date" in cols else "DATE(i.created_at)"


# business day → [start, end) timestamps using cutoff hour (0–23)
def _business_window(business_date: str, cutoff_hour: int):
    day = datetime.strptime(business_date, "%Y-%m-%d")
    start = day.replace(hour=cutoff_hour, minute=0, second=0)
    end = start + timedelta(days=1)
    return start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")


def _invoice_date_expr(cur) -> str:
    """
    Returns the SQL expression to use for the invoice 'date' filter,
    preferring i.invoice_date if it exists, otherwise DATE(i.created_at).
    """
    cur.execute("PRAGMA table_info(invoices)")
    cols = [row[1] for row in cur.fetchall()]
    return "DATE(i.invoice_date)" if "invoice_date" in cols else "DATE(i.created_at)"


# ---------- Core data fetcher (used by both PDF + on-screen + Excel) ----------
def get_z_report_data(business_date: str, cutoff_hour: int = 0):
    """
    Returns:
      {
        "window": {"start_ts","end_ts","cutoff_hour"},
        "sales":  {"invoice_count","gross_sales","tax_collected",
                   "net_by_rate":[...], "vat_by_rate":[...], "gross_by_rate":[...],
                   "totals":{"net","vat","gross"}},
        "tenders": {"by_method":[{"method","payments_total","refunds_total"}...],
                    "net_total": float},
      }
    """
    import sqlite3

    conn = _connect()
    cur = conn.cursor()

    start_ts, end_ts = _business_window(business_date, int(cutoff_hour))
    date_expr = _invoice_date_expr(cur)

    # --- Sales: invoices for that business date (invoice_date) ---
    # Gross sales = SUM(final_amount), VAT by rate discount-aware, Net = Gross - VAT
    # Discount-aware per-rate rollup
    vat_sql = f"""
    WITH inv AS (
      SELECT i.invoice_id, COALESCE(i.discount,0) AS disc_pct, COALESCE(i.final_amount,0) AS final_amount
      FROM invoices i
      WHERE {date_expr} = :D
        AND i.invoice_type = 'INVOICE'
        AND COALESCE(i.revenue_eligible,1) = 1
    ),
    lines AS (
      SELECT ii.invoice_id,
             COALESCE(
               CAST(ROUND(ii.vat_pct * 100.0, 0) AS INTEGER),
               CASE ii.vat_flag WHEN 'B' THEN 5 WHEN 'C' THEN 19 ELSE 0 END, 0
             ) AS vat_rate,
             SUM(ii.total_price - COALESCE(ii.vat_amount,0)) AS net_pre,
             SUM(COALESCE(ii.vat_amount,0))                  AS vat_pre,
             SUM(ii.total_price)                             AS gross_pre
      FROM invoice_items ii
      JOIN inv ON inv.invoice_id = ii.invoice_id
      GROUP BY ii.invoice_id, vat_rate
    ),
    adj AS (
      SELECT l.vat_rate,
             (l.net_pre   * (1.0 - inv.disc_pct/100.0)) AS net_adj,
             (l.vat_pre   * (1.0 - inv.disc_pct/100.0)) AS vat_adj,
             (l.gross_pre * (1.0 - inv.disc_pct/100.0)) AS gross_adj
      FROM lines l JOIN inv ON inv.invoice_id = l.invoice_id
    )
    SELECT vat_rate,
           ROUND(SUM(net_adj),  2) AS net_excl_vat,
           ROUND(SUM(vat_adj),  2) AS vat_amount,
           ROUND(SUM(gross_adj),2) AS gross_incl_vat
    FROM adj
    GROUP BY vat_rate
    ORDER BY vat_rate;
    """
    cur.execute(vat_sql, {"D": business_date})
    rows = cur.fetchall()
    by_rate = [
        {
            "rate": int(r[0]),
            "net": float(r[1] or 0.0),
            "vat": float(r[2] or 0.0),
            "gross": float(r[3] or 0.0),
        }
        for r in rows
    ]
    totals = {
        "net": round(sum(r["net"] for r in by_rate), 2),
        "vat": round(sum(r["vat"] for r in by_rate), 2),
        "gross": round(sum(r["gross"] for r in by_rate), 2),
    }

    # Invoice count & gross sales = SUM(final_amount) for the day
    cur.execute(
        f"""
        SELECT COUNT(*), ROUND(SUM(COALESCE(final_amount,0)), 2)
          FROM invoices i
         WHERE {date_expr} = :D
           AND i.invoice_type = 'INVOICE'
           AND COALESCE(i.revenue_eligible,1) = 1
    """,
        {"D": business_date},
    )
    inv_count, gross_sales = cur.fetchone() or (0, 0.0)

    # --- Tenders & refunds: payments within the business window ---
    # We currently don't track refunds; if you ever store them as negative payments,
    # this split keeps them separate.
    tender_sql = """
    SELECT
      COALESCE(payment_method,'Other') AS method,
      ROUND(SUM(CASE WHEN amount_paid >= 0 THEN amount_paid ELSE 0 END), 2) AS payments_total,
      ROUND(ABS(SUM(CASE WHEN amount_paid <  0 THEN amount_paid ELSE 0 END)), 2) AS refunds_total
    FROM payment_history
    WHERE DATETIME(payment_date) >= DATETIME(:START)
      AND DATETIME(payment_date) <  DATETIME(:END)
    GROUP BY payment_method
    ORDER BY method;
    """
    cur.execute(tender_sql, {"START": start_ts, "END": end_ts})
    tenders = [
        {
            "method": m or "Other",
            "payments_total": float(p or 0.0),
            "refunds_total": float(r or 0.0),
        }
        for (m, p, r) in cur.fetchall()
    ]
    net_total = round(sum(t["payments_total"] - t["refunds_total"] for t in tenders), 2)

    conn.close()

    return {
        "window": {
            "start_ts": start_ts,
            "end_ts": end_ts,
            "cutoff_hour": int(cutoff_hour),
        },
        "sales": {
            "invoice_count": int(inv_count or 0),
            "gross_sales": float(gross_sales or 0.0),
            "tax_collected": float(totals["vat"]),
            "by_rate": by_rate,
            "totals": totals,
        },
        "tenders": {"by_method": tenders, "net_total": net_total},
    }


# ---------- PDF export ----------
# â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â
# Helper: find logo in a few common spots
# â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â
def _find_logo_path():
    import os

    here = os.path.dirname(__file__)
    candidates = [
        os.path.join(here, "pet_wellness_logo.png"),
        os.path.join(here, "logo.png"),
        os.path.join(os.path.dirname(here), "pet_wellness_logo.png"),
        os.path.join(os.path.dirname(here), "logo.png"),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


# â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â
# Export Z-Report → PDF (with header, logo, section titles, table headers)
# â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â
def export_z_report_pdf(parent, business_date: str, cutoff_hour: int = 0):
    from datetime import datetime

    from PySide6.QtWidgets import QFileDialog, QMessageBox
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    # Pull business-day data
    data = get_z_report_data(business_date, cutoff_hour)
    sales = data["sales"]
    tend = data["tenders"]
    win = data["window"]

    # Save-as path
    default_name = f"Z_Report_{business_date}.pdf"
    path, _ = QFileDialog.getSaveFileName(
        parent, "Save Z-Report (PDF)", default_name, "PDF Files (*.pdf)"
    )
    if not path:
        return

    # Styles
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle(
        "H1", parent=styles["Heading1"], fontSize=16, leading=18, spaceAfter=8
    )
    H2 = ParagraphStyle(
        "H2", parent=styles["Heading2"], fontSize=13, leading=15, spaceAfter=6
    )
    SMALL = ParagraphStyle("SMALL", parent=styles["Normal"], fontSize=9, leading=11)
    NORMAL = styles["Normal"]

    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=15 * mm,
    )

    elems = []

    # â â  Header with clinic info (left) + logo (right)
    clinic_left = [
        Paragraph("<b>PET WELLNESS VETS</b>", H1),
        Paragraph("Kyriakou Adamou no.2, Shop 2&3, 8220", SMALL),
        Paragraph("Tel: 99941186", SMALL),
        Paragraph("Email: contact@petwellnessvets.com", SMALL),
        Paragraph("VAT / Tax ID: 60118644C", SMALL),
    ]
    logo_fp = _find_logo_path()
    logo_cell = [Image(logo_fp, width=35 * mm, height=18 * mm)] if logo_fp else []
    header_tbl = Table([[clinic_left, logo_cell]], colWidths=[120 * mm, 49 * mm])
    header_tbl.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ]
        )
    )
    elems += [header_tbl, Spacer(1, 6 * mm)]

    # Title + business window
    title_tbl = Table(
        [
            [
                Paragraph("<b>Z-REPORT</b>", styles["Title"]),
                Paragraph(
                    f"<b>Business Date:</b> {business_date}<br/>"
                    f"<b>Cutoff Hour:</b> {cutoff_hour}<br/>"
                    f"<b>Window:</b> {win['start_ts']} → {win['end_ts']}",
                    SMALL,
                ),
            ]
        ],
        colWidths=[70 * mm, 99 * mm],
    )
    title_tbl.setStyle(TableStyle([("ALIGN", (1, 0), (1, 0), "RIGHT")]))
    elems += [title_tbl, Spacer(1, 6 * mm)]

    # â â  Tenders & Refunds
    elems += [Paragraph("TENDERS & REFUNDS (Payments by payment_date)", H2)]
    tender_rows = [["Method", "Payments total", "Refunds total"]]
    order = [
        "CASH",
        "CARD",
        "BANK_TRANSFER",
        "ONLINE",
        "CHEQUE",
        "GIFT_CARD",
        "STORE_CREDIT",
    ]
    by_name = {t["method"]: t for t in tend["by_method"]}
    for name in order:
        row = by_name.get(name, {"payments_total": 0.0, "refunds_total": 0.0})
        tender_rows.append(
            [name, f"€{row['payments_total']:.2f}", f"€{row['refunds_total']:.2f}"]
        )
    # Net total line
    tender_rows.append(["Net total", f"€{tend['net_total']:.2f}", ""])
    t_tbl = Table(tender_rows, colWidths=[70 * mm, 45 * mm, 35 * mm])
    t_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),  # header
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c8c8c8")),
                ("ALIGN", (1, 1), (-1, -2), "RIGHT"),
                ("ALIGN", (1, -1), (1, -1), "RIGHT"),
                ("FONTNAME", (0, -1), (0, -1), "Helvetica-Bold"),
            ]
        )
    )
    elems += [t_tbl, Spacer(1, 8 * mm)]

    # â â  Sales
    elems += [Paragraph("SALES (Invoices by invoice_date)", H2)]
    sales_lines = [
        [
            "Invoices count (FINAL/POSTED/PAID):",
            str(int(sales.get("invoice_count", 0))),
        ],
        ["Gross sales:", f"€{sales['totals']['gross']:.2f}"],
        ["VAT / TAX BREAKDOWN (FINAL & POSTED; invoice_date in window):", ""],
        ["Tax collected:", f"€{sales['totals']['vat']:.2f}"],
    ]
    s_tbl = Table(sales_lines, colWidths=[120 * mm, 49 * mm])
    s_tbl.setStyle(
        TableStyle(
            [
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    elems += [s_tbl, Spacer(1, 4 * mm)]

    # VAT by rate table
    vat_rows = [["Rate", "Net taxable", "VAT amount", "Gross (Net+VAT)"]]
    for r in sales["by_rate"]:
        vat_rows.append(
            [
                f"{int(r['rate'])}%",
                f"€{r['net']:.2f}",
                f"€{r['vat']:.2f}",
                f"€{r['gross']:.2f}",
            ]
        )
    vat_rows.append(
        [
            "TOTAL",
            f"€{sales['totals']['net']:.2f}",
            f"€{sales['totals']['vat']:.2f}",
            f"€{sales['totals']['gross']:.2f}",
        ]
    )
    v_tbl = Table(vat_rows, colWidths=[25 * mm, 45 * mm, 40 * mm, 49 * mm])
    v_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),  # header
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c8c8c8")),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]
        )
    )
    elems += [v_tbl, Spacer(1, 10 * mm)]

    # Footer
    elems += [Paragraph(f"Generated at: {datetime.now():%Y-%m-%d %H:%M:%S}", SMALL)]

    # Build PDF
    doc.build(elems)

    QMessageBox.information(parent, "Z-Report", f"Saved: {path}")


# ---------- Excel export (single-sheet Z_REPORT) ----------
def export_z_report_excel(parent, business_date: str, cutoff_hour: int = 0):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from PySide6.QtWidgets import QFileDialog, QMessageBox

    data = get_z_report_data(business_date, cutoff_hour)
    sales = data["sales"]
    tend = data["tenders"]
    win = data["window"]

    # Static clinic info (edit to your needs or read from a settings table)
    clinic = {
        "Clinic/Company": "S&K Pet Wellness Vet LTD",
        "Address Line 1": "2 Kyriakou Adamou",
        "Address Line 2": "Shop 2&3",
        "City/Postal": "Paphos, Chloraka 8220",
        "Country": "Cyprus",
        "Phone": "35726910310",
        "Email": "contact@petwellnessvest.com",
        "Website": "www.petwellnessvets.com",
        "VAT / Tax ID": "60118644C",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Z_REPORT"
    B = Font(bold=True)
    R = Alignment(horizontal="right")
    C = Alignment(horizontal="center")

    r = 1
    # Left: clinic block
    for k, v in clinic.items():
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        r += 1

    ws.cell(r, 1, "Business Date")
    ws.cell(r, 2, business_date)
    r += 1
    ws.cell(r, 1, "Cutoff Hour (0–23)")
    ws.cell(r, 2, cutoff_hour)
    r += 1
    ws.cell(r, 1, "Start Timestamp")
    ws.cell(r, 2, win["start_ts"])
    r += 1
    ws.cell(r, 1, "End Timestamp")
    ws.cell(r, 2, win["end_ts"])
    r += 2

    # Right: tenders & refunds table header
    ws.cell(1, 5, "TENDERS & REFUNDS (Payments by payment_date)").font = B
    ws.cell(2, 5, "Method").font = B
    ws.cell(2, 6, "Payments total").font = B
    ws.cell(2, 7, "Refunds total").font = B

    # Fill tenders
    rr = 3
    order = [
        "CASH",
        "CARD",
        "BANK_TRANSFER",
        "ONLINE",
        "CHEQUE",
        "GIFT_CARD",
        "STORE_CREDIT",
    ]
    by_name = {t["method"]: t for t in tend["by_method"]}
    for name in order:
        row = by_name.get(name, {"payments_total": 0.0, "refunds_total": 0.0})
        ws.cell(rr, 5, name)
        ws.cell(rr, 6, row["payments_total"]).alignment = R
        ws.cell(rr, 7, row["refunds_total"]).alignment = R
        rr += 1

    # Net total line
    ws.cell(rr + 1, 5, "Net total").font = B
    ws.cell(rr + 1, 6, tend["net_total"]).alignment = R
    # spacer row
    r = max(r, rr + 3)

    # SALES section
    ws.cell(r, 1, "SALES (Invoices by invoice_date)").font = B
    r += 2
    ws.cell(r, 1, "Invoices count (FINAL/POSTED/PAID)")
    ws.cell(r, 2, int(sales["invoice_count"]))
    r += 1
    ws.cell(r, 1, "Gross sales")
    ws.cell(r, 2, float(sales["totals"]["gross"])).alignment = R
    r += 1
    ws.cell(
        r, 1, "VAT / TAX BREAKDOWN (FINAL & POSTED; invoice_date in window)"
    ).font = B
    r += 1
    ws.cell(r, 1, "Tax collected")
    ws.cell(r, 2, float(sales["totals"]["vat"])).alignment = R
    r += 1

    # VAT breakdown table (Rate/Net/VAT/Gross)
    ws.cell(r, 1, "Rate").font = B
    ws.cell(r, 2, "Net taxable").font = B
    ws.cell(r, 3, "VAT amount").font = B
    ws.cell(r, 4, "Gross (Net+VAT)").font = B
    r += 1
    for row in sales["by_rate"]:
        ws.cell(r, 1, f"{row['rate']}%")
        ws.cell(r, 2, row["net"]).alignment = R
        ws.cell(r, 3, row["vat"]).alignment = R
        ws.cell(r, 4, row["gross"]).alignment = R
        r += 1
    ws.cell(r, 1, "TOTAL").font = B
    ws.cell(r, 2, sales["totals"]["net"]).alignment = R
    ws.cell(r, 2).font = B
    ws.cell(r, 3, sales["totals"]["vat"]).alignment = R
    ws.cell(r, 3).font = B
    ws.cell(r, 4, sales["totals"]["gross"]).alignment = R
    ws.cell(r, 4).font = B
    r += 2

    # Cash drawer section (placeholders; zero until you add drawer events/refunds tables)
    zeros = [
        "Opening float (sum of sessions starting in window)",
        "Cash in (drawer events)",
        "Cash out + payouts (drawer events)",
        "Net cash tenders (payments - refunds)",
        "Expected closing cash",
        "Counted cash (sum of sessions ending in window)",
        "Over / Short (Counted - Expected)",
    ]
    vals = [0, 0, 0, tend["net_total"], 0, 0, 0]
    for label, val in zip(zeros, vals):
        ws.cell(r, 1, label)
        ws.cell(r, 2, float(val)).alignment = R
        r += 1

    # widths
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 24
    for col in ("E", "F", "G"):
        ws.column_dimensions[col].width = 18

    # save
    default = f"Z_Report_{business_date}.xlsx"
    path, _ = QFileDialog.getSaveFileName(
        parent, "Save Z-Report (Excel)", default, "Excel Files (*.xlsx)"
    )
    if not path:
        return
    wb.save(path)
    QMessageBox.information(parent, "Z-Report", f"Saved: {path}")


# ---------- Simple in-app screen ----------
class ZReportWidget(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)

        # â â  Top controls: Business Date + Cutoff Hour + actions â â â â â â â â â â â â â â â â â
        ctl = QHBoxLayout()
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setCalendarPopup(True)

        self.cutoff_spin = QSpinBox()
        self.cutoff_spin.setRange(0, 23)
        self.cutoff_spin.setValue(0)  # default cutoff 00:00–24:00
        self.cutoff_spin.setToolTip("Cutoff hour for the business day (0–23).")

        self.refresh_btn = QPushButton("Refresh")
        self.pdf_btn = QPushButton("Export PDF")
        self.xlsx_btn = QPushButton("Export Excel")

        ctl.addWidget(QLabel("Business date:"))
        ctl.addWidget(self.date_edit)
        ctl.addSpacing(12)
        ctl.addWidget(QLabel("Cutoff:"))
        ctl.addWidget(self.cutoff_spin)
        ctl.addStretch(1)
        ctl.addWidget(self.refresh_btn)
        ctl.addWidget(self.pdf_btn)
        ctl.addWidget(self.xlsx_btn)
        layout.addLayout(ctl)

        # â â  Summary strip â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â
        self.summary_label = QLabel("")
        self.summary_label.setWordWrap(True)
        layout.addWidget(self.summary_label)

        # â â  VAT by rate table â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â
        layout.addWidget(QLabel("VAT by rate"))
        self.vat_tbl = QTableWidget(0, 4)
        self.vat_tbl.setHorizontalHeaderLabels(
            ["VAT %", "Net excl. VAT", "VAT", "Gross incl. VAT"]
        )
        self.vat_tbl.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.vat_tbl.verticalHeader().setVisible(False)
        layout.addWidget(self.vat_tbl)

        # â â  Payments by method table â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â
        layout.addWidget(QLabel("Payments received in business window (by method)"))
        self.pay_tbl = QTableWidget(0, 3)
        self.pay_tbl.setHorizontalHeaderLabels(["Method", "Amount", "Count"])
        self.pay_tbl.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.pay_tbl.verticalHeader().setVisible(False)
        layout.addWidget(self.pay_tbl)

        # â â  Wiring â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â â
        self.refresh_btn.clicked.connect(self.refresh)
        # Your export_* methods should read self.cutoff_spin.value()
        self.pdf_btn.clicked.connect(self.export_pdf)
        self.xlsx_btn.clicked.connect(self.export_xlsx)

        # Auto-refresh when date or cutoff changes
        self.date_edit.dateChanged.connect(lambda _d: self.refresh())
        self.cutoff_spin.valueChanged.connect(lambda _v: self.refresh())

        # Initial load
        self.refresh()

    # inside reports.py, class ZReportWidget …

    def _date_str(self) -> str:
        return self.date_edit.date().toString("yyyy-MM-dd")

    def refresh(self):
        # pull data with cutoff
        business_date = self._date_str()
        cutoff = self.cutoff_spin.value()
        data = get_z_report_data(business_date, cutoff)

        sales = data.get("sales", {})
        tenders = data.get("tenders", {})
        totals = sales.get("totals", {})
        by_rate = sales.get("by_rate", [])
        by_method = tenders.get("by_method", [])
        net_tenders = float(tenders.get("net_total", 0.0))

        # â â  summary strip
        self.summary_label.setText(
            "Invoices: <b>{cnt}</b> | "
            "Gross: <b>€{gross:.2f}</b> | "
            "VAT: <b>€{vat:.2f}</b> | "
            "Net: <b>€{net:.2f}</b> | "
            "Tenders net: <b>€{tnet:.2f}</b>".format(
                cnt=int(sales.get("invoice_count", 0)),
                gross=float(totals.get("gross", 0.0)),
                vat=float(totals.get("vat", 0.0)),
                net=float(totals.get("net", 0.0)),
                tnet=net_tenders,
            )
        )

        # â â  VAT by rate table
        self.vat_tbl.setRowCount(0)
        for r in by_rate:
            row = self.vat_tbl.rowCount()
            self.vat_tbl.insertRow(row)
            self.vat_tbl.setItem(row, 0, QTableWidgetItem(f"{int(r.get('rate', 0))}%"))
            self.vat_tbl.setItem(
                row, 1, QTableWidgetItem(f"{float(r.get('net', 0.0)):.2f}")
            )
            self.vat_tbl.setItem(
                row, 2, QTableWidgetItem(f"{float(r.get('vat', 0.0)):.2f}")
            )
            self.vat_tbl.setItem(
                row, 3, QTableWidgetItem(f"{float(r.get('gross', 0.0)):.2f}")
            )

        # â â  Payments by method table (payments & refunds columns)
        # ensure headers match the new structure
        self.pay_tbl.setColumnCount(3)
        self.pay_tbl.setHorizontalHeaderLabels(
            ["Method", "Payments total", "Refunds total"]
        )
        self.pay_tbl.setRowCount(0)

        # show common methods in a consistent order
        order = [
            "CASH",
            "CARD",
            "BANK_TRANSFER",
            "ONLINE",
            "CHEQUE",
            "GIFT_CARD",
            "STORE_CREDIT",
        ]
        by_name = {
            m.get("method"): "{:.2f}|{:.2f}".format(
                float(m.get("payments_total", 0.0)), float(m.get("refunds_total", 0.0))
            )
            for m in by_method
        }

        for method in order:
            pay_ref = by_name.get(method, "0.00|0.00").split("|")
            payments = pay_ref[0]
            refunds = pay_ref[1]
            row = self.pay_tbl.rowCount()
            self.pay_tbl.insertRow(row)
            self.pay_tbl.setItem(row, 0, QTableWidgetItem(method))
            self.pay_tbl.setItem(row, 1, QTableWidgetItem(f"{float(payments):.2f}"))
            self.pay_tbl.setItem(row, 2, QTableWidgetItem(f"{float(refunds):.2f}"))

    def export_pdf(self):
        export_z_report_pdf(self, self._date_str(), self.cutoff_spin.value())

    def export_xlsx(self):
        export_z_report_excel(self, self._date_str(), self.cutoff_spin.value())
